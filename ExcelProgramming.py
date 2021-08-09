import openpyxl
import pathlib
import csv


"""
book = openpyxl.Workbook()
sheet_title = "Python_Excel.xlsx"
book.save(sheet_title)
"""








def print_header():
    ws["A1"] = "AAA"
    ws["A2"] = "BBB"
    ws["A3"] = "CCC"
    ws["A4"] = "DDD"
    ws["A5"] = "EEE"












book = "Python_Excel.xlsx"



wb = openpyxl.load_workbook(book)
ws = wb["Sheet"]


print_header()


ws["B6"] = 1200
ws.cell(3, 4).value = 42


wb.save(book)








